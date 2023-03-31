import sys
from sys import argv
from openpyxl import load_workbook
from variables import WORKBOOK_DEFAULT, SHEET_RANGE_MAX_DEFAULT as max_range, SHEET_RANGE_MIN_DEFAULT as min_range


""" Module for load data from XLSX file """


def row_gen(load_file=WORKBOOK_DEFAULT, sheet_range_min=min_range, sheet_range_max=max_range):
    """ Генератор строки из таблицы с данными """

    workbook = load_workbook(load_file, read_only=True, data_only=True)
    sheets = workbook.sheetnames
    for sheet in sheets:
        worksheet = workbook[sheet]
        rows_from_table = (worksheet[sheet_range_min:sheet_range_max])
        i_row = ((cell.value for cell in row) for row in rows_from_table)
        # получаем дату из ячеек 'C3' и 'D3'
        date1, date2 = worksheet['C3'].value, worksheet['D3'].value
    return i_row, date1, date2


if __name__ == '__main__':
    """ Для запуска через коммандную строку:
        - Обычный запуск: python xls_load.py
        - Запуск с параметрами: python xls_load.py base_data.xlsx """
    param = argv
    try:
        var = param[1]
        print(f' Параметры запуска: {param}')
        rows, date1, date2 = row_gen(argv[1], 'A4', 'J8')
    except IndexError:
        rows, date1, date2 = row_gen(WORKBOOK_DEFAULT, 'B4', 'J999')

    print(f' Размер генератора _ {sys.getsizeof(rows)}')
    for row in rows:
        print(f'{list(row)} _ размер строки _ {sys.getsizeof(row)}')
    try:
        print(list(next(rows)))
    except StopIteration:
        print(f' Данных больше нет ')
